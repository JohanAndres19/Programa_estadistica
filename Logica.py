from openpyxl import *
from tkinter import *
from tkinter import messagebox
from ventana import *
from tkinter import filedialog
import numpy as np 

class  modelo:
    def __init__(self):
        self.raiz = Raiz(self)
        self.ventanainicial = None         
        self.CambioVentana(0)
        self.resultado=[]
        self.raiz._tk.mainloop()
         
    def CambioVentana (self,opcion):
        if opcion== 0:
            self.raiz.Get_VentanaInicio().Set_visible(True)
        elif opcion == 1:
            if self.raiz.Get_VentanaInicio().Isvisibe()==True:
                self.raiz.Get_VentanaInicio().pack_forget()
                self.raiz.Get_VentanaInicio().Set_visible(False)
                self.raiz.Get_VentanaIngresar().Set_visible(True)
                self.raiz.Get_VentanaIngresar().Organizar()
            else:
                self.raiz.Get_VentanaIngresar().pack_forget()
                self.raiz.Get_VentanaIngresar().ventana.place_forget()
                self.raiz.Get_VentanaIngresar().botonAtras.place_forget()
                self.raiz.Get_VentanaInicio().master.geometry("420x200")
                self.raiz.Get_VentanaInicio().Set_visible(True)
                self.raiz.Get_VentanaIngresar().Set_visible(False)
                self.raiz.Get_VentanaInicio().pack()
        elif opcion == 2:
            if self.raiz.Get_VentanaOperaciones().Isvisibe():
                self.raiz.Get_VentanaOperaciones().pack_forget()
                self.raiz.Get_VentanaInicio().master.geometry("420x200")
                self.raiz.Get_VentanaInicio().pack()
        elif opcion ==3:
            if self.raiz.Get_VentanaIngresar().Isvisibe()==True:
                self.raiz.Get_VentanaIngresar().Ocultar()
                if len(self.raiz.Get_VentanaOperaciones().Get_Datos1())!=0:
                  self.raiz.Get_VentanaOperaciones().destroy()
                  self.raiz.Set_VentanaOperaciones(VentanaOperaciones(self.raiz._tk,self))
                self.raiz.Get_VentanaOperaciones().Set_Datos1(self.raiz.Get_VentanaIngresar().ventana.Get_Variable1())
                self.raiz.Get_VentanaOperaciones().Set_Datos2(self.raiz.Get_VentanaIngresar().ventana.Get_Variable2())
                self.raiz.Get_VentanaOperaciones().Componentes()
                self.raiz.Get_VentanaOperaciones().Set_visible(True)
                self.raiz.Get_VentanaOperaciones().Get_botonatras().config(command=self.raiz.Get_VentanaOperaciones().controlador.Boton_Atras)
                self.raiz.Get_VentanaIngresar().Set_visible(False)
                self.raiz.Get_VentanaOperaciones().master.geometry("550x450")

                self.raiz.Get_VentanaOperaciones().pack(fill="y", expand=1)        
            else:
                self.raiz.Get_VentanaOperaciones().Set_visible(False)
                self.raiz.Get_VentanaOperaciones().paneldatos.grid_forget()
                self.raiz.Get_VentanaOperaciones().pack_forget()
                self.raiz.Get_VentanaIngresar().Set_visible(True)
                self.raiz.Get_VentanaIngresar().Organizar()
        elif opcion == 4:
            if  self.raiz.Get_VentanaOperaciones().Isvisibe()==True:  
                self.raiz.Get_VentanaOperaciones().pack_forget()
                self.raiz.Get_VentanaOperaciones().Set_visible(False)
                self.raiz.Get_VentanaResultado().Set_visible(True)
                self.raiz.Get_VentanaResultado().Componentes(self.resultado)
            else:
                self.raiz.Get_VentanaOperaciones().Set_visible(True)
                self.raiz.Get_VentanaResultado().Set_visible(False)
                self.raiz.Get_VentanaResultado().pack_forget()
                self.raiz.Get_VentanaOperaciones().pack(fill="y", expand=1)
        elif opcion ==5:
            if  self.raiz.Get_VentanaOperaciones().Isvisibe()==True:  
                self.raiz.Get_VentanaOperaciones().pack_forget()
                self.raiz.Get_VentanaOperaciones().Set_visible(False)
                self.raiz.Get_VentanaResultado().Set_visible(True)
                self.raiz.Get_VentanaResultado().ComponentesMann(self.resultado)   
            else:
                self.raiz.Get_VentanaOperaciones().Set_visible(True)
                self.raiz.Get_VentanaResultado().Set_visible(False)
                self.raiz.Get_VentanaResultado().pack_forget()
                self.raiz.Get_VentanaOperaciones().pack(fill="y", expand=1)
    
    def Tablaingresardatos(self):
        try:
            if self.raiz.Get_VentanaIngresar().activo:
                self.raiz.Get_VentanaIngresar().valor1=int(self.raiz.Get_VentanaIngresar().Valorvariable1.get())
                self.raiz.Get_VentanaIngresar().valor2=int(self.raiz.Get_VentanaIngresar().Valorvariable2.get())            
                if self.raiz.Get_VentanaIngresar().valor1 in range(1,11) and self.raiz.Get_VentanaIngresar().valor2 in range(1,11):
                    self.raiz.Get_VentanaIngresar().ventana.ConstruirTabla(self.raiz.Get_VentanaIngresar().valor1,self.raiz.Get_VentanaIngresar().valor2)
                    self.raiz.Get_VentanaIngresar().Is_activoboton1(0)
                    self.raiz.Get_VentanaIngresar().Is_activoboton2(0)
                    self.raiz.Get_VentanaIngresar().activo=False
                    self.raiz.Get_VentanaIngresar().botonborrar1.config(state="active")
                    self.raiz.Get_VentanaIngresar().botonborrar2.config(state="active")
                else:
                    messagebox.showwarning("Advertencia","Por favor ingrese valores validos"+"\n y que esten en el rango de 1 a 10")        
            else:
                if self.raiz.Get_VentanaIngresar().activoVariable1==True and self.raiz.Get_VentanaIngresar().activoVariable2==False:
                    aux= int(self.raiz.Get_VentanaIngresar().Valorvariable1.get())
                    if aux in range(1,11):
                        if aux > self.raiz.Get_VentanaIngresar().valor1:
                            self.raiz.Get_VentanaIngresar().ventana.Set_Agregar(True) 
                        elif aux <self.raiz.Get_VentanaIngresar().valor1:
                            self.raiz.Get_VentanaIngresar().ventana.Set_Agregar(False) 
                        self.raiz.Get_VentanaIngresar().valor1=aux
                        self.raiz.Get_VentanaIngresar().ventana.ConstruirTabla(aux,0)
                        self.raiz.Get_VentanaIngresar().Is_activoboton1(0)
                    else:
                        raise Exception('')        
                elif self.raiz.Get_VentanaIngresar().activoVariable2==True and self.raiz.Get_VentanaIngresar().activoVariable1==False:
                    aux= int(self.raiz.Get_VentanaIngresar().Valorvariable2.get())
                    if aux in range(1,11):
                        if aux > self.raiz.Get_VentanaIngresar().valor2:
                            self.raiz.Get_VentanaIngresar().ventana.Set_Agregar(True) 
                        elif aux <self.raiz.Get_VentanaIngresar().valor2:
                            self.raiz.Get_VentanaIngresar().ventana.Set_Agregar(False) 
                        self.raiz.Get_VentanaIngresar().valor2=aux
                        self.raiz.Get_VentanaIngresar().ventana.ConstruirTabla(0,aux)
                        self.raiz.Get_VentanaIngresar().Is_activoboton2(0)    
                    else:
                        raise Exception('')   
                elif self.raiz.Get_VentanaIngresar().activoVariable2==True and self.raiz.Get_VentanaIngresar().activoVariable1==True:
                    aux1 =int(self.raiz.Get_VentanaIngresar().Valorvariable1.get())
                    aux2=int(self.raiz.Get_VentanaIngresar().Valorvariable2.get()) 
                    if  aux1 in range(1,11) and aux2 in range(1,11):
                        if aux1 >self.raiz.Get_VentanaIngresar().valor1 and aux2>self.raiz.Get_VentanaIngresar().valor2:
                            self.raiz.Get_VentanaIngresar().ventana.Set_opcion((1,1))
                        elif aux1 <self.raiz.Get_VentanaIngresar().valor1 and aux2<self.raiz.Get_VentanaIngresar().valor2:
                            self.raiz.Get_VentanaIngresar().ventana.Set_opcion((0,0))
                            self.raiz.Get_VentanaIngresar().Is_activoboton2(0)
                        elif  aux1 <self.raiz.Get_VentanaIngresar().valor1 and aux2>self.raiz.Get_VentanaIngresar().valor2:  
                            self.raiz.Get_VentanaIngresar().ventana.Set_opcion((0,1))
                        elif   aux1 >self.raiz.Get_VentanaIngresar().valor1 and aux2<self.raiz.Get_VentanaIngresar().valor2:
                            self.raiz.Get_VentanaIngresar().ventana.Set_opcion((1,0))
                        self.raiz.Get_VentanaIngresar().valor1=aux1
                        self.raiz.Get_VentanaIngresar().valor2=aux2     
                        self.raiz.Get_VentanaIngresar().ventana.ConstruirTabla(aux1,aux2)
                        self.raiz.Get_VentanaIngresar().Is_activoboton1(0)
                        self.raiz.Get_VentanaIngresar().Is_activoboton2(0)
                    else:
                       raise Exception              
        except: 
            messagebox.showwarning("Advertencia","Por favor ingrese valores validos"+"\n y que esten en el rango de 1 a 10")

    def onclickCargarexcelinterface(self):
        self.filename = filedialog.askopenfilename(title="Archivo", filetypes=(
            ("xlsx files", "*.xlsx"), ("all files", "*.*")))
        if self.filename!="":
            datos=CargarDatos(self.filename)
            datos.Cargar()
            if datos.Get_Cargado():
                if len(self.raiz.Get_VentanaOperaciones().Get_Datos1())!=0:
                  self.raiz.Get_VentanaOperaciones().destroy()
                  self.raiz.Set_VentanaOperaciones(VentanaOperaciones(self.raiz._tk,self))
                self.raiz.Get_VentanaInicio().pack_forget()
                self.raiz.Get_VentanaInicio().Set_visible(False)
                self.raiz.Get_VentanaOperaciones().master.geometry("550x450")
                self.raiz.Get_VentanaOperaciones().Set_visible(True)
                self.raiz.Get_VentanaOperaciones().Set_Datos1(datos.Get_Valores1())
                self.raiz.Get_VentanaOperaciones().Set_Datos2(datos.Get_Valores2())
                self.raiz.Get_VentanaOperaciones().Componentes()
                self.raiz.Get_VentanaOperaciones().Get_botonatras().config(command=self.raiz.Get_VentanaOperaciones().controlador.Boton_AtrasI)
        else:
            messagebox.showinfo("Ventana","Seleccione un archivo por favor")

    def CalcularWilcoxon(self):
        calcularw=None
        if self.raiz.Get_VentanaOperaciones().Get_combohipoalter().get()!="seleccione una opcion" and self.raiz.Get_VentanaOperaciones().Get_comboopcione().get()!=" ":
            if len(self.raiz.Get_VentanaOperaciones().Get_Datos1()) <=len(self.raiz.Get_VentanaOperaciones().Get_Datos2()):
                calcularw=Wilcoxon(self.raiz.Get_VentanaOperaciones().Get_Datos1(),self.raiz.Get_VentanaOperaciones().Get_Datos2(),self.raiz.Get_VentanaOperaciones().Get_combohipoalter().get(),self.raiz.Get_VentanaOperaciones().Get_comboopcione().get())
            else:
                calcularw=Wilcoxon(self.raiz.Get_VentanaOperaciones().Get_Datos2(),self.raiz.Get_VentanaOperaciones().Get_Datos1(),self.raiz.Get_VentanaOperaciones().Get_combohipoalter().get(),self.raiz.Get_VentanaOperaciones().Get_comboopcione().get())
            calcularw.Algoritmo() 
            if  calcularw.Get_Solucion()!="": 
                if len(self.raiz.Get_VentanaResultado().Get_listarespuesta())!=0:
                    self.raiz.Get_VentanaResultado().destroy()
                    self.raiz.Set_VentanaResultado(VentanaResultados(self.raiz._tk,self))
                self.resultado=calcularw.Get_Resultado()   
                self.CambioVentana(4)
        else:
            messagebox.showwarning("ventana","Porfavor elija una opcion valida")       

    def CalcularMann_Whitney(self):
        calcularMann=None 
        try:
            if self.raiz.Get_VentanaOperaciones().Get_combohipoalter().get()!="seleccione una opcion" and self.raiz.Get_VentanaOperaciones().Get_JtextDato().get()!=" ":
                alpha= float(self.raiz.Get_VentanaOperaciones().Get_JtextDato().get())
                if  alpha>=0 and alpha<=0.6:
                    if len(self.raiz.Get_VentanaOperaciones().Get_Datos1()) <=len(self.raiz.Get_VentanaOperaciones().Get_Datos2()):
                        calcularMann= Mann_Whitney(self.raiz.Get_VentanaOperaciones().Get_Datos1(),self.raiz.Get_VentanaOperaciones().Get_Datos2(),alpha,self.raiz.Get_VentanaOperaciones().Get_combohipoalter().get())
                    else:
                        calcularMann= Mann_Whitney(self.raiz.Get_VentanaOperaciones().Get_Datos2(),self.raiz.Get_VentanaOperaciones().Get_Datos1(),alpha,self.raiz.Get_VentanaOperaciones().Get_combohipoalter().get())           
                    if len(self.raiz.Get_VentanaResultado().Get_listarespuesta())!=0:
                        self.raiz.Get_VentanaResultado().destroy()
                        self.raiz.Set_VentanaResultado(VentanaResultados(self.raiz._tk,self))
                    self.resultado=calcularMann.Algoritmo()
                    if self.resultado!=None:
                        if self.resultado[len(self.resultado)-1]!="":
                            self.CambioVentana(5)
                else:
                    raise Exception
        except Exception:
            messagebox.showinfo("Ventana","Por favor ingrese un alpha que se encuentre entre 0 y 0.6 ")          
        except:
          messagebox.showwarning("Advertencia","Ingrese valores validos por favor y recuerdo usar un punto ( . ) para las cifrass decimales")       

class CargarDatos:

    _alpha=None     
    def __init__(self,url):
        try:
            self.url=url
            self.archivo = load_workbook(self.url)
            self.lista =self.archivo.sheetnames
            self.hojadetrabajo = self.archivo[self.lista[0]]
            self.lista1=[]
            self.lista2=[]
        except:
            messagebox.showerror("Error","No se pudo cargar el archivo por favor intente otra vez")    
    
    def Cargar(self):
        try:
            for fila in self.hojadetrabajo.rows:
                auxiliar=0
                for celda in fila:
                    if celda.value != None and celda.row !=1:
                        if celda.column ==1:
                            aux=celda.value
                            if type(aux) is str: 
                                if aux.find('.')==-1:
                                    self.lista1.append(int(aux))
                                else:
                                    self.lista1.append(float(aux))   
                            else:
                                self.lista1.append(aux)
                            auxiliar=celda.row            
                        elif celda.column == 2:
                            aux=celda.value
                            if type(aux) is str:
                                if aux.find('.')==-1:
                                    self.lista2.append(int(aux))
                                else:
                                    self.lista2.append(float(aux))
                            else:
                                self.lista2.append(aux)
                            auxiliar=celda.row                    
                        else:
                            auxiliar=celda.row
                            break; 
                if auxiliar>11:
                   raise Exception('')                    
            self.cargado =True
        except Exception:
            self.cargado=False
            messagebox.showwarning("Advertencia","Los datos son maximo 10 para cada variable"+" \nrevise el archivo e intente de nuevo")            
        except :
            self.cargado=False
            messagebox.showwarning("Advertencia","Datos erroneos revise el formato")  

    @classmethod
    def CargarTablasWilcoxon(cls,n1,n2,alpha,opcion):
        try:      
            if opcion=="μ1<μ2" or opcion=="μ1>μ2":
                return cls.CargarUnaCola(n1,n2,alpha)
            else:
                return cls.CargarDosColas(n1,n2,alpha)
        except Exception:
            messagebox.showwarning("Advertencia","El valor de la variable 2 debe ser mayor o igual a 3")

    @classmethod
    def CargarUnaCola(cls,n1,n2,alpha):
        tabla= load_workbook("./resources/Tablawilcoxon.xlsx")
        hojas = tabla.sheetnames
        if n2>=3:
                if  alpha=="0.001" :
                    tabla1=tabla[hojas[0]]
                    columna=1
                    if n2>=6:
                        columna=n2-6            
                        return tabla1.cell(row=(3+n1),column=(2+columna)).value
                    else:
                        raise Exception
                elif  alpha == "0.01":
                    tabla2=tabla[hojas[1]]
                    if n2>=5:
                        columna= n2-5
                        return tabla2.cell(row=3+n1,column=(2+columna)).value
                    else:
                        raise Exception
                elif alpha ==  "0.025" :
                    tabla3=tabla[hojas[2]]
                    if n2>=4:
                        columna= n2-4
                        return tabla3.cell(row=(3+n1),column=(2+columna)).value
                    else:
                        raise Exception
                elif alpha ==  "0.05":
                    tabla4=tabla[hojas[3]]
                    if n2>=3:
                        columna=n2-3
                        return tabla4.cell(row=(2+n1),column=(2+columna)).value
                    else:
                        raise Exception
        else:
            raise Exception
    
    @classmethod
    def CargarDosColas(cls,n1,n2,alpha):      
        tabla= load_workbook("./resources/Tablawilcoxon.xlsx")
        hojas = tabla.sheetnames  
        if n2>=3:
            if  alpha== "0.002":
                tabla1=tabla[hojas[0]]
                columna=1
                if n2>=6:
                    columna=n2-6
                    
                    return tabla1.cell(row=(3+n1),column=(2+columna)).value
                else:
                    raise Exception
            elif alpha =="0.02" :
                tabla2=tabla[hojas[1]]
                if n2>=5:
                    columna= n2-5
                    return tabla2.cell(row=2+n1,column=(2+columna)).value
                else:
                    raise Exception
            elif alpha=="0.05":
                tabla3=tabla[hojas[2]]
                if n2>=4:
                    columna= n2-4
                    return tabla3.cell(row=(2+n1),column=(2+columna)).value
                else:
                    raise Exception
            elif alpha=="0.1":
                tabla4=tabla[hojas[3]]
                if n2>=3:
                    columna=n2-3
                    return tabla4.cell(row=(2+n1),column=(2+columna)).value
                else:
                    raise Exception
        else:
                raise Exception

    @classmethod 
    def CargarTablaMann(cls,n1,n2,alpha):
        try:
            tabla= load_workbook("./resources/mann_whyn.xlsx")
            lista_nombre= tabla.sheetnames
            listavalores=[]
            if n2>=3:
                hojadetrabajo = tabla[lista_nombre[n2-3]]    
                for i in range(3,hojadetrabajo.max_row+1):
                    aux=hojadetrabajo.cell(row=i,column=n1+1).value
                    if aux !=None:
                     listavalores.append(float(aux))
                valor_apro= np.array(listavalores)
                alpha_aprox=valor_apro.flat[np.abs(valor_apro-alpha).argmin()]
                fila=listavalores.index(alpha_aprox)
                fila+=3
                cls.Set_alpha_aprox(alpha_aprox)
                return hojadetrabajo.cell(row=fila,column=1).value
        except:
            messagebox.showerror("Error","No se pudo cargar el valor de la tabla")   
    @classmethod
    def Set_alpha_aprox(cls,valor):
        cls._alpha=valor

    @classmethod
    def Get_alpha_aprox(cls):
        return cls._alpha

    def Get_Valores1(self):
        return self.lista1    

    def Get_Valores2(self):
        return self.lista2        

    def Get_Cargado(self):
        return self.cargado            

class Wilcoxon :
    def __init__(self,variable1,variable2, opcion,alpha):
        self.numeros= []
        self.rango= []
        self.lista1=variable1
        self.lista2=variable2
        self.opcion=opcion
        self.alpha=alpha
        self.resultado=[]
        self.solucion=""

    def Algoritmo(self):
       self.u=CargarDatos.CargarTablasWilcoxon(len(self.lista1),len(self.lista2),self.alpha,self.opcion)
       if self.u!=None:
          self.Algoritmo_parte2()
          
       
    def Algoritmo_parte2(self):
        self.calcular_metodo()
        valor=""
        nom_asociado=()
        if self.opcion=="μ1<μ2": 
            if self.u1<=self.u:
                valor="Hay diferencia suficiente para  rechazar la hipotesis nula"
            else:
                valor=" No hay diferencia suficiente para rechaza la hipotesis nula"       
            nom_asociado=(self.u1,"U1")
            self.resultado.append(nom_asociado)
        elif self.opcion=="μ1>μ2":
            if self.u2<=self.u:
                valor="Hay diferencia suficiente para  rechazar la hipotesis nula"
            else:
                valor=" No hay diferencia suficiente para rechaza la hipotesis nula"
            nom_asociado=(self.u2,"U2")
            self.resultado.append(nom_asociado)
        else:
            if self.u1<self.u2:
                if self.u1<=self.u:
                    valor="Hay diferencia suficiente para  rechazar la hipotesis nula"
                else:
                    valor=" No hay diferencia suficiente para rechaza la hipotesis nula"
                nom_asociado=(self.u1,"U1")
                self.resultado.append(nom_asociado)
            elif self.u2<self.u1:
                if self.u2<=self.u:
                    valor="Hay diferencia suficiente para  rechazar la hipotesis nula"
                else:
                    valor=" No hay diferencia suficiente para rechaza la hipotesis nula"
                nom_asociado=(self.u2,"U2")
                self.resultado.append(nom_asociado)
            else:
                if self.u2<=self.u:
                    valor="Hay diferencia suficiente para  rechazar la hipotesis nula"
                else:
                    valor=" No hay diferencia suficiente para rechaza la hipotesis nula" 
                nom_asociado=(self.u2,"U2")
                self.resultado.append(nom_asociado)
        self.resultado.append(self.opcion)
        self.resultado.append(self.w1)
        self.resultado.append(self.w2)
        self.resultado.append(self.u)
        self.resultado.append(valor)
        self.solucion=valor
    
    def calcular_metodo(self):
        suma=0
        for i in self.lista1:
            self.numeros.append(i)
        
        for j in self.lista2:
            self.numeros.append(j)
        
        self.MetodoQuicSort()
        self.Rango() 
        
        for i in self.lista1:
            suma += self.rango[self.numeros.index(i)]
        self.w1=suma
        self.w2 =(((len(self.lista1)+len(self.lista2))*(len(self.lista1)+len(self.lista2)+1))/2)-suma
        self.u1 = self.w1-(len(self.lista1)*(len(self.lista1)+1))/2
        self.u2= self.w2 -(len(self.lista2)*(len(self.lista2)+1))/2
    
    def Get_Resultado(self):
        return self.resultado

    def Get_Solucion(self):
        return self.solucion
    
    def MetodoQuicSort(self):
        pivote=0
        for i in range(1,len(self.numeros)):
            j=i
            while j>0 and self.numeros[j-1]>self.numeros[j]:
                pivote =self.numeros[j-1]
                self.numeros[j-1]=self.numeros[j]
                self.numeros[j]= pivote
                j-=1

    def Rango(self):
        aux=0
        for i in range(len(self.numeros)):
            if i < len(self.numeros)-1:
                if self.numeros[i] ==self.numeros[i+1]:
                    valormedio=((1+aux)+(2+aux))/2
                    self.rango.append(valormedio)
                    self.rango.append(valormedio)
                    aux+=3
                    self.rango.append(aux)
                elif aux == len(self.numeros):
                    break
                else:
                    aux+=1
                    self.rango.append(aux)
  
class  Mann_Whitney:
    def __init__(self,variable1,vaeiable2,alpha,opcion) :
        self.lista1=variable1
        self.lista2=vaeiable2             
        self.numeros=[]
        self.rango=[]  
        self.alpha=alpha  
        self.opcion=opcion
        self.resultado=[]

    def Algoritmo(self):
        try:
            self.calcular_metodo()
            solucion=""
            self.u =(len(self.lista1)*len(self.lista2))+((len(self.lista1)*(len(self.lista1)+1))/2)-self.w1
            self.resultado.append(self.u)
            self.resultado.append(self.w1)
            if self.opcion == "μ1<μ2" or self.opcion=="μ1>μ2":
                self.u0=CargarDatos.CargarTablaMann(len(self.lista1),len(self.lista2),self.alpha)
                if self.u0 !=None:
                    self.resultado.append(self.u0)
                    if self.opcion =="μ1<μ2":
                        if self.u>= (len(self.lista1)*len(self.lista2)-self.u0):
                            solucion="Se rechaza la hipotesis nula por haber una diferencia significativa"  
                        else:
                            solucion="No se puede rechaza la hipotesis Nula  al no haber una diferencia significativa"
                        aux=(len(self.lista1)*len(self.lista2)-self.u0)
                        self.resultado.append(aux)
                    else:
                        if self.u<=self.u0:
                            solucion="Se rechaza la hipotesis nula por haber una diferencia significativa"  
                        else:
                            solucion="No se puede rechaza la hipotesis Nula  al no haber una diferencia significativa"
                    self.resultado.append(CargarDatos.Get_alpha_aprox())
                else:
                    raise Exception()
            else:
                alpha=self.alpha/2
                self.u0=CargarDatos.CargarTablaMann(len(self.lista1),len(self.lista2),alpha)
                if self.u0 != None:
                    self.resultado.append(self.u0)
                    if self.u>= (len(self.lista1)*len(self.lista2)-self.u0) or self.u<=self.u0:
                        solucion="Se rechaza la hipotesis nula por haber una diferencia significativa"  
                    else:
                        solucion="No se puede rechaza la hipotesis Nula  al no haber una diferencia significativa"
                    aux=(len(self.lista1)*len(self.lista2)-self.u0)
                    self.resultado.append(aux)
                    self.resultado.append(CargarDatos.Get_alpha_aprox())
                else:
                    raise Exception()   
            self.resultado.append(self.alpha)
            self.resultado.append(solucion)
            self.resultado.append(self.opcion)
            return self.resultado
        except Exception :
            messagebox.showwarning("Advertencia","Por favor Verificar que al menos una de las variables tenga minimo como minimo 3 elementos")

    def calcular_metodo(self):
        suma=0
        for i in self.lista1:
            self.numeros.append(i)
        
        for j in self.lista2:
            self.numeros.append(j)
        
        self.MetodoQuicSort()
        self.Rango() 

        for i in self.lista1:
            suma += self.rango[self.numeros.index(i)]
        self.w1=suma

    def MetodoQuicSort(self):
        pivote=0
        for i in range(1,len(self.numeros)):
            j=i
            while j>0 and self.numeros[j-1]>self.numeros[j]:
                pivote =self.numeros[j-1]
                self.numeros[j-1]=self.numeros[j]
                self.numeros[j]= pivote
                j-=1

    def Rango(self):
        aux=0
        for i in range(len(self.numeros)):
            if i < len(self.numeros)-1:
                if self.numeros[i] ==self.numeros[i+1]:
                    valormedio=((1+aux)+(2+aux))/2
                    self.rango.append(valormedio)
                    self.rango.append(valormedio)
                    aux+=3
                    self.rango.append(aux)
                elif aux == len(self.numeros):
                    break
                else:
                    aux+=1
                    self.rango.append(aux)
  